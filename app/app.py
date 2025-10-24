import re

# -*- coding: utf-8 -*-
import io, json, shutil, csv, datetime as dt
from pathlib import Path
from flask import Flask, render_template, request, send_file, make_response, jsonify, Response
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from zipfile import BadZipFile
from openpyxl.utils.exceptions import InvalidFileException
from copy import copy as _cpy_style

def _uniq_files(files):
    seen=set(); uniq=[]
    for f in files:
        key = (f.filename, getattr(f, "content_length", None))
        if key not in seen:
            seen.add(key); uniq.append(f)
    return uniq

# === Label toggle (single column) & Max load writers ===
SEP_CHARS = r"／/|｜・,、\-"
DIGITS = r"0-9０-９"

def _normalize_spaces2(s: str) -> str:
    s = s.replace("\u3000", " ")
    return re.sub(r"\s+", " ", s).strip()

def _strip_label_text_single(text: str, keep: str) -> str:
    """単一セル内の「増車/減車」表記から不要側を安全に除去。"""
    if not isinstance(text, str) or not text:
        return text
    t = _normalize_spaces2(text)
    remove_label = "減車" if keep == "増車" else "増車"
    unit = r"(?:台|輌|両|件|回|枚)?"
    pat = rf"""
        (?:^|[\s{SEP_CHARS}])
        {remove_label}\s*
        (?:[{DIGITS}]+)?\s*{unit}?
        (?:\([^\)]*\))?
        (?=$|[\s{SEP_CHARS}])
    """
    t = re.sub(pat, " ", t, flags=re.VERBOSE)
    t = re.sub(rf"\s*([{SEP_CHARS}])\s*", r"\1", t)
    t = re.sub(rf"[{SEP_CHARS}]{{2,}}", "/", t)
    t = t.strip(" /|｜・,、-")
    t = _normalize_spaces2(t)
    return t

def apply_label_toggle_single(xlsx: Path, modes, sheet_index: int = 0,
                              col_inc="C", col_dec="M", start_row=21, end_row=25):
    """21..25行で対象列のみ：増車→C列、減車→M列。B21へB5を転記。"""
    wb = _safe_load_wb(xlsx, data_only=False)
    ws = wb.worksheets[sheet_index]

    rows = list(range(start_row, end_row + 1))
    modes = list(modes)
    limit = min(len(rows), len(modes))

    for i in range(limit):
        row = rows[i]
        keep = "増車" if str(modes[i]).lower().startswith("inc") or modes[i] == "増車" else "減車"
        col = col_inc if keep == "増車" else col_dec
        cell = ws[f"{col}{row}"]
        if getattr(cell, "data_type", None) == "f":
            continue
        val = cell.value
        if isinstance(val, str) and val.strip():
            cell.value = _strip_label_text_single(val, keep)

    # B21 に拠点名（B5と同じ）を転記
    try:
        b5 = ws["B5"].value
        if b5:
            ws["B21"].value = b5
    except Exception:
        pass

    try:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass

    wb.save(xlsx)
    wb.close()

def _get_letter_max_load(file_storage, mode: str) -> int | None:
    """連絡書から最大積載量を取得。増車:U23／減車:AQ23（先頭シート）。整数にして返す。"""
    try:
        if hasattr(file_storage, "seek"):
            file_storage.seek(0)
        by = file_storage.read() if hasattr(file_storage, "read") else None
        if by is None:
            return None
        wb = load_workbook(io.BytesIO(by), data_only=True)
        ws = wb.worksheets[0]
        addr = "U23" if (str(mode).lower().startswith("inc") or mode == "増車") else "AQ23"
        v = ws[addr].value
        try:
            import re as _re2
            if isinstance(v, str):
                m = _re2.search(r"([0-9０-９,\.]+)", v)
                v = m.group(1) if m else ""
                v = v.replace(",", "").replace("，", "")
            if v in (None, ""):
                return None
            return int(float(v))
        finally:
            wb.close()
    except Exception:
        return None

def apply_max_loads(xlsx: Path, modes, inc_files, dec_files, sheet_index: int = 0,
                    start_row=21, end_row=25, dest_col="F"):
    """F21..F25に「<整数> kg」を書く。増車→inc_files、減車→dec_files を順番に対応。"""
    wb = _safe_load_wb(xlsx, data_only=False)
    ws = wb.worksheets[sheet_index]

    rows = list(range(start_row, end_row + 1))
    modes = list(modes)
    limit = min(len(rows), len(modes))

    ii = 0; dd = 0
    for i in range(limit):
        row = rows[i]
        mode = modes[i]
        if str(mode).lower().startswith("inc") or mode == "増車":
            if ii >= len(inc_files): 
                continue
            fileobj = inc_files[ii]; ii += 1
        else:
            if dd >= len(dec_files):
                continue
            fileobj = dec_files[dd]; dd += 1

        val = _get_letter_max_load(fileobj, mode)
        try:
            if hasattr(fileobj, "seek"):
                fileobj.seek(0)
        except Exception:
            pass
        if val is not None:
            ws[f"{dest_col}{row}"].value = f"{val} kg"

    try:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass

    wb.save(xlsx)
    wb.close()


# ==== Safe helpers injected ====
def _resolve_mode_key(mapping:dict, mode:str)->str:
    aliases = {
        "increase": ["increase","inc","増車","増"],
        "decrease": ["decrease","dec","減車","減"]
    }
    for k in aliases.get(mode, [mode]):
        if k in mapping: return k
    # try loose compare ignoring spaces
    norm = lambda s: "".join(str(s).split())
    for k in list(mapping.keys()):
        if norm(k) in [norm(x) for x in aliases.get(mode, [mode])]:
            return k
    raise ValueError("mapping.yaml に '増車/減車'（increase/decrease）に対応する設定が見つかりません。")

def fail(msg: str, status: int = 400):
    from flask import make_response
    import json
    html = f'<script>alert({json.dumps(msg)});history.back();</script>'
    r = make_response(html, status)
    r.headers["Content-Type"] = "text/html; charset=utf-8"
    return r

def is_valid_xlsx(path: Path) -> tuple[bool, str]:
    try:
        if not path.exists():
            return False, "not found"
        if path.stat().st_size < 200:
            return False, "too small"
        with path.open("rb") as f:
            if f.read(4) != b"PK\x03\x04":  # xlsx=zip signature
                return False, "bad signature"
        wb = load_workbook(path, read_only=True, data_only=True)
        wb.close()
        return True, ""
    except Exception as e:
        return False, str(e)
    
# UIセレクト値 → 正規の列名（ヘッダ）     ※counter.json のヘッダ名に合わせる
WEIGHT_UI_TO_HEADER = {
    "to_2t": "2.0トンまで",
    "2t_long": "2.0トンロング",
    "over_2t_long_to_7_5t": "2.0トンロング超～7.5トンまで",
    "over_7_5t": "7.5トンを超えるもの",
}

def _norm_key(s: str) -> str:
    """スペース/全角スペース/中黒/ドット等を除いて比較用に正規化"""
    if s is None: return ""
    s = str(s)
    s = s.replace("　", " ").replace("・", "")
    s = re.sub(r"[ \t\r\n.]", "", s)   # 空白とドットを削る
    return s

def _find_weight_db_key(rec: dict, header_label: str) -> str | None:
    """
    レコード内のキーから、ヘッダ名（例: '2.0トンまで'）に対応するキーを探す。
    例: '2 . 0  トンまで' のような空白混じりも拾う。
    """
    target = _norm_key(header_label)
    # 1) 完全一致があれば最優先
    if header_label in rec:
        return header_label
    # 2) 正規化一致
    for k in rec.keys():
        if _norm_key(k) == target:
            return k
    return None

def update_db_weight_inplace(db_flat: dict, row_idx: int, ui_value: str, delta: int) -> bool:
    """
    UIセレクト値（to_2t 等）をDB列にマップし、その列が存在すれば増減。
    列が見つからない場合は何もしない（DB構造は変更しない）。
    """
    rec = db_flat["records"][row_idx]
    header_label = WEIGHT_UI_TO_HEADER.get(ui_value or "", "")
    if not header_label:
        return False
    db_key = _find_weight_db_key(rec, header_label)
    if not db_key:
        return False
    now = rec.get(db_key, 0)
    try:
        now = int(str(now).strip())
    except Exception:
        now = 0
    rec[db_key] = now + int(delta)
    save_db_raw(db_flat)  # ← 既存の保存関数を利用
    return True    


def _ensure_mode_keys(mapping:dict)->dict:
    # make a shallow copy and ensure canonical English keys exist
    mp = dict(mapping)
    # promote JP keys to EN
    if "increase" not in mp:
        for k in ("増車","増"):
            if k in mp: mp["increase"] = mp[k]; break
    if "decrease" not in mp:
        for k in ("減車","減"):
            if k in mp: mp["decrease"] = mp[k]; break
    # if still missing one side, mirror the other so動作が止まらない（値は同一）
    if "increase" not in mp and "decrease" in mp:
        mp["increase"] = mp["decrease"]
    if "decrease" not in mp and "increase" in mp:
        mp["decrease"] = mp["increase"]
    return mp

def _safe_load_wb(file_or_bytes, data_only: bool = True):
    try:
        if hasattr(file_or_bytes, "read"):
            return load_workbook(io.BytesIO(file_or_bytes.read()), data_only=data_only)
        else:
            return load_workbook(file_or_bytes, data_only=data_only)
    except (BadZipFile, InvalidFileException) as e:
        raise ValueError("連絡書は.xlsx形式のExcelファイルのみ対応です。ファイルを確認してください。") from e
# ==== end helpers ====
from zipfile import BadZipFile
from openpyxl.utils.exceptions import InvalidFileException
import yaml
app = Flask(__name__)
ROOT = Path(__file__).resolve().parent
SRC = ROOT / "templates_src"
OUTPUT = ROOT / "outputs"
DATA = ROOT / "data"
OUTPUT.mkdir(exist_ok=True); DATA.mkdir(exist_ok=True)
TPL_NAME = "申請書共通.xlsx"

def js_alert(msg: str, status: int = 400):
    html = f"""<script>alert({json.dumps(msg)});history.back();</script>"""
    # 4xxで返してXHR側を確実に失敗扱いにする
    r = make_response(html, status)
    r.headers["Content-Type"] = "text/html; charset=utf-8"
    return r

def norm_spaces(s:str)->str:
    if s is None: return ""
    for z in ("\u00A0","\u200B","\u200C","\u200D"): s=s.replace(z,"")
    return s.replace(" ","").replace("\t","").replace("\u3000","").strip()

def load_mapping()->dict:
    with (ROOT/"mapping.yaml").open("r",encoding="utf-8") as f: return yaml.safe_load(f)

def read_contact_values(file_or_bytes, mode:str, mapping:dict)->dict:
    src = mapping[_resolve_mode_key(mapping, mode)]
    if hasattr(file_or_bytes,"read"): wb = _safe_load_wb(file_or_bytes)
    else: wb = _safe_load_wb(file_or_bytes)
    ws = wb.worksheets[0]
    def at(a): 
        if not a: return ""
        v = ws[a].value; return str(v).strip() if v is not None else ""
    out = {"user_name":at(src.get("user_name")),"office_name":at(src.get("office_name")),
           "era":at(src.get("era")),"year":at(src.get("year")),"capacity":at(src.get("capacity"))}
    wb.close(); return out

def parse_office_display(name:str):
    name=(name or "").strip()
    if not name: return "","",""
    if "本社" in name and "営業所" in name: return "本社営業所","", "本社"
    if name.endswith("営業所"): loc=name[:-3]; return f"{loc}営業所", loc, "営業所"
    if name.endswith("支店"): loc=name[:-2]; return f"{loc}支店", loc, "支店"
    return name, name, ""

# DB I/O
def _db_path()->Path: return DATA/"counter.json"

def load_db_raw()->dict:
    p=_db_path()
    if not p.exists(): return {"headers":[], "records":[]}
    with p.open("r",encoding="utf-8") as f: raw=json.load(f)
    if isinstance(raw,dict) and "area_table" in raw and isinstance(raw["area_table"],dict):
        at=raw["area_table"]; return {"headers":at.get("headers",[]), "records":at.get("records",[])}
    return {"headers":raw.get("headers",[]), "records":raw.get("records",[])}

def save_db_raw(db_flat:dict)->None:
    p=_db_path()
    base={}
    if p.exists():
        try:
            with p.open("r",encoding="utf-8") as f: base=json.load(f) or {}
        except Exception: base={}
    base["area_table"]={"headers":db_flat.get("headers",[]),"records":db_flat.get("records",[])}
    with p.open("w",encoding="utf-8") as f: json.dump(base,f,ensure_ascii=False,indent=2)

def find_company_index(db_flat:dict, company:str, office_display:str)->int:
    company_n=norm_spaces(company)
    disp, loc, divi = parse_office_display(office_display)
    loc_n=norm_spaces(loc)
    headers=db_flat.get("headers",[]); recs=db_flat.get("records",[])
    def val(rec,key): return rec.get(key,"") if isinstance(rec,dict) else rec[headers.index(key)]
    for i,r in enumerate(recs):
        comp=norm_spaces(val(r,"会社名")); div=str(val(r,"本社/営業所/支店") or "").strip()
        locv=norm_spaces(str(val(r,"営業所所在地") or ""))
        if comp==company_n and (div==(divi or "本社")) and ((div=="本社" and locv=="") or (div!="本社" and locv==loc_n)):
            return i
    return -1

def get_type_counts_from_row(rec:dict)->dict:
    def num(v): 
        try: return int(float(str(v).strip()))
        except: return 0
    return {"普通":num(rec.get("普通",0)), "小型":num(rec.get("小型",0)),
            "けん引":num(rec.get("けん引",0)), "被けん引":num(rec.get("被けん引",0))}

def update_db_counts_inplace(db_flat:dict, idx:int, car_type:str, delta:int)->dict:
    rec=db_flat["records"][idx]; now=get_type_counts_from_row(rec)
    rec[car_type]=int(now.get(car_type,0))+int(delta)
    total=sum(get_type_counts_from_row(rec).values())
    if "合計" in rec: rec["合計"]=total
    save_db_raw(db_flat); return get_type_counts_from_row(rec)

def put_old_new_counts(xlsx: Path, office_disp: str, old_counts: dict, new_counts: dict):
    wb = _safe_load_wb(xlsx, data_only=False)   # ★ここをFalseにする
    ws = wb.worksheets[0]

    # 入力セルのみを 5〜8 行でクリア（I/R 列や 9 行の式は触れない）
    input_cols = ["B", "D", "E", "G", "H", "K", "N", "O", "Q"]
    for r in range(5, 9):  # 5..8
        for c in input_cols:
            ws[f"{c}{r}"] = None

    r = 5
    ws[f"B{r}"] = office_disp
    ws[f"K{r}"] = int(old_counts.get("普通", 0))
    ws[f"N{r}"] = int(old_counts.get("小型", 0))
    ws[f"O{r}"] = int(old_counts.get("けん引", 0))
    ws[f"Q{r}"] = int(old_counts.get("被けん引", 0))
    ws[f"D{r}"] = int(new_counts.get("普通", 0))
    ws[f"E{r}"] = int(new_counts.get("小型", 0))
    ws[f"G{r}"] = int(new_counts.get("けん引", 0))
    ws[f"H{r}"] = int(new_counts.get("被けん引", 0))

    # 3) Excelで開いたときに式を必ず再計算させる（念のため）
    try:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass

    wb.save(xlsx)
    wb.close()

def put_weight_totals_to_sheet(xlsx_path: str, rec: dict) -> None:
    """
    申請書の L30～L33 に、DB（rec）の重量別カウントを転記する。
    B30～B33 のラベル文字と rec のキーを表記ゆれ吸収で突き合わせる。
    """
    # _safe_load_wb と _find_weight_db_key は既存のものを利用
    wb = _safe_load_wb(xlsx_path, data_only=False)
    ws = wb.worksheets[0]   # 先頭シート固定

    # 30～33行を走査：B列のラベル → DBキー解決 → L列に数値
    for r in (30, 31, 32, 33):
        label = str(ws[f"B{r}"].value or "").strip()
        if not label:
            continue
        # 表記ゆれ吸収で rec の実キーを取得（既存ヘルパ）
        key = _find_weight_db_key(rec, label)
        if not key:
            continue
        # 整数化して L 列へ
        try:
            val = int(str(rec.get(key, 0)).strip() or "0")
        except Exception:
            val = 0
        ws[f"L{r}"].value = val

    wb.save(xlsx_path)
    wb.close()

# ==== app.py 追記（importの下あたり。既存ヘルパ群と同じ階層）====
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HONSHI_INC_PATH = SRC / "申請書増車1.xlsx"
HONSHI_DEC_PATH = SRC / "申請書減車1.xlsx"

def _copy_sheet_content(src_ws, dst_ws):
    """シート間コピー（値・基本書式・列幅/行高・結合・フリーズ等）。画像/図形/コメントは対象外。"""
    # 値とスタイル
    for row in src_ws.iter_rows():
        for c in row:

            d = dst_ws.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                try:
                    d.font        = _cpy_style(c.font)
                    d.border      = _cpy_style(c.border)
                    d.fill        = _cpy_style(c.fill)
                    d.number_format = c.number_format  # これは文字列なのでコピー不要
                    d.protection  = _cpy_style(c.protection)
                    d.alignment   = _cpy_style(c.alignment)
                except TypeError:
        # 万一Hash不可のオブジェクトが混じっても落とさない
                    pass
    # 列幅
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
    # 行高
    for r_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[r_idx].height = dim.height
    # 結合
    if src_ws.merged_cells.ranges:
        for rng in list(src_ws.merged_cells.ranges):
            dst_ws.merge_cells(str(rng))
    # フリーズ
    dst_ws.freeze_panes = src_ws.freeze_panes
    # ページ設定（必要最小限）
    try:
        dst_ws.page_setup.orientation = src_ws.page_setup.orientation
        dst_ws.page_setup.paperSize = src_ws.page_setup.paperSize
        dst_ws.page_margins.left = src_ws.page_margins.left
        dst_ws.page_margins.right = src_ws.page_margins.right
        dst_ws.page_margins.top = src_ws.page_margins.top
        dst_ws.page_margins.bottom = src_ws.page_margins.bottom
    except Exception:
        pass

def prepend_honshi_sheet(out_path: Path, prefer: str):
    """
    out_path: 生成済みの別紙ファイル（すでに内容転記済み）
    prefer: "increase" か "decrease"
    既存のシートを「別紙」にリネームし、先頭に「本紙」を追加して保存。
    """
    # どちらの本紙を使うか
    honshi_path = HONSHI_INC_PATH if prefer == "increase" else HONSHI_DEC_PATH
    # 既存出力
    wb = load_workbook(out_path)
    # 既存シート名調整（シート2）
    try:
        wb.worksheets[0].title = "別紙"
    except Exception:
        pass
    # 本紙の読み込み
    src_wb = load_workbook(honshi_path, data_only=False)
    src_ws = src_wb.worksheets[0]
    # シート1を作ってコピー
    dst_ws = wb.create_sheet(title="本紙", index=0)
    _copy_sheet_content(src_ws, dst_ws)
    src_wb.close()
    # 保存
    wb.save(out_path)
    wb.close()


@app.get("/")
def index(): return render_template("index.html")

@app.get("/db")
def db_get(): return jsonify(load_db_raw())

@app.post("/db/save")
def db_save():
    try: payload=request.get_json(force=True)
    except Exception: return "invalid json", 400
    if not isinstance(payload,dict) or "records" not in payload: return "bad payload", 400
    headers=payload.get("headers",[]); recs=payload.get("records",[])
    if not headers:
        keys=[]; 
        for r in recs:
            if isinstance(r,dict):
                for k in r.keys():
                    if k not in keys: keys.append(k)
        headers=keys
    num_cols={"普通","小型","けん引","被けん引","合計"}
    normed=[]
    for r in recs:
        row=dict(r) if isinstance(r,dict) else {}
        for k in list(row.keys()):
            v=row[k]
            if k in num_cols:
                try: row[k]=int(float(str(v).strip())) if str(v).strip()!="" else 0
                except Exception: row[k]=0
        normed.append(row)
    save_db_raw({"headers":headers,"records":normed}); return "ok", 200

@app.get("/db/csv")
def db_csv():
    flat = load_db_raw()
    headers = flat.get("headers", [])
    recs = flat.get("records", [])

    def generate():
        yield ",".join(headers) + "\n"
        for r in recs:
            row = [str((r.get(h, "") if isinstance(r, dict) else "")) for h in headers]
            yield ",".join(row) + "\n"

    return Response(
        generate(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=counter.csv"},
    )

# === Block-based label removal & model year writers (C-block/M-block, H-block/R-block) ===
def apply_label_toggle_blocked(xlsx: Path, modes, sheet_index: int = 0,
                               start_row=21, end_row=25, col_block=("C","M")):
    """1〜5件目→C列、6〜10件目→M列のみ処理（増車/減車に依存しない）。"""
    wb = _safe_load_wb(xlsx, data_only=False)
    ws = wb.worksheets[sheet_index]

    rows = list(range(start_row, end_row + 1))  # 21..25
    modes = list(modes)[:10]  # 最大10件
    for i, mode in enumerate(modes):
        row = rows[i % len(rows)]
        # ブロック判定：0..4 → C、5..9 → M
        col = col_block[0] if i < 5 else col_block[1]
        keep = "増車" if str(mode).lower().startswith("inc") or mode == "増車" else "減車"
        cell = ws[f"{col}{row}"]
        if getattr(cell, "data_type", None) == "f":
            continue
        val = cell.value
        if isinstance(val, str) and val.strip():
            cell.value = _strip_label_text_single(val, keep)

    # B21 に B5 の営業所名を転記
    try:
        b5 = ws["B5"].value
        if b5:
            ws["B21"].value = b5
    except Exception:
        pass

    try:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass

    wb.save(xlsx)
    wb.close()

def _clean_era(v) -> str | None:
    s = "" if v is None else str(v).strip().upper()
    s = s.replace("Ｒ","R").replace("Ｈ","H")
    m = re.search(r"[RH]", s)
    return m.group(0) if m else None

def _clean_year_num(v) -> int | None:
    if v is None:
        return None
    s = str(v)
    m = re.search(r"([0-9０-９,\.]+)", s)
    if not m:
        return None
    s = m.group(1).replace(",","").replace("，","")
    try:
        return int(float(s))
    except Exception:
        return None

def _get_letter_model_year(file_storage, mode: str) -> str | None:
    """増車: R19/U19、減車: AM19/AQ19 → 'H27' などに整形。"""
    try:
        if hasattr(fileobj := file_storage, "seek"):
            fileobj.seek(0)
        by = fileobj.read() if hasattr(fileobj, "read") else None
        if by is None:
            return None
        wb = load_workbook(io.BytesIO(by), data_only=True)
        ws = wb.worksheets[0]
        if str(mode).lower().startswith("inc") or mode == "増車":
            era = _clean_era(ws["R19"].value)
            num = _clean_year_num(ws["U19"].value)
        else:
            era = _clean_era(ws["AM19"].value)
            num = _clean_year_num(ws["AQ19"].value)
        wb.close()
        if era and (num is not None):
            return f"{era}{num}"
        return None
    except Exception:
        return None

def apply_model_years(xlsx: Path, modes, inc_files, dec_files, sheet_index: int = 0,
                      start_row=21, end_row=25, dest_cols=("H","R")):
    """1〜5件目→H21..H25、6〜10件目→R21..R25 に 'H27' 等を記入。"""
    wb = _safe_load_wb(xlsx, data_only=False)
    ws = wb.worksheets[sheet_index]

    rows = list(range(start_row, end_row + 1))  # 21..25
    modes = list(modes)[:10]

    ii = dd = 0
    for i, mode in enumerate(modes):
        row = rows[i % len(rows)]
        dest_col = dest_cols[0] if i < 5 else dest_cols[1]

        if str(mode).lower().startswith("inc") or mode == "増車":
            if ii >= len(inc_files):
                continue
            fobj = inc_files[ii]; ii += 1
        else:
            if dd >= len(dec_files):
                continue
            fobj = dec_files[dd]; dd += 1

        val = _get_letter_model_year(fobj, mode)
        try:
            if hasattr(fobj, "seek"):
                fobj.seek(0)
        except Exception:
            pass
        if val:
            ws[f"{dest_col}{row}"].value = val

    try:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass

    wb.save(xlsx)
    wb.close()

@app.post("/process")
def process():
    inc_files = [f for f in request.files.getlist("increase_files") if getattr(f, "filename", "")]
    dec_files = [f for f in request.files.getlist("decrease_files") if getattr(f, "filename", "")]
    
    # ファイル別の選択値を取得
    inc_car_types = []
    inc_weight_classes = []
    dec_car_types = []
    dec_weight_classes = []
    
    # 増車ファイルの選択値を取得
    for i in range(len(inc_files)):
        car_type = request.form.get(f"inc_car_type_{i}", "普通").strip()
        weight_class = request.form.get(f"inc_weight_class_{i}", "to_2t").strip()
        inc_car_types.append(car_type)
        inc_weight_classes.append(weight_class)
    
    # 減車ファイルの選択値を取得
    for i in range(len(dec_files)):
        car_type = request.form.get(f"dec_car_type_{i}", "普通").strip()
        weight_class = request.form.get(f"dec_weight_class_{i}", "to_2t").strip()
        dec_car_types.append(car_type)
        dec_weight_classes.append(weight_class)
    inc_files = _uniq_files(inc_files)
    dec_files = _uniq_files(dec_files)
    if not inc_files and not dec_files:
        return js_alert("増車・減車どちらかのファイルをアップロードしてください。", status=400)

    MAP = _ensure_mode_keys(load_mapping())
    metas = []
    for f in inc_files:
        metas.append(("increase", read_contact_values(f, "increase", MAP)))
    for f in dec_files:
        metas.append(("decrease", read_contact_values(f, "decrease", MAP)))

    companies = {m[1]["user_name"] for m in metas}
    offices = {m[1]["office_name"] for m in metas}
    if len(companies) != 1 or len(offices) != 1:
        return js_alert("アップロード内で会社名/営業所名が混在しています。1社1拠点で実行してください。", status=400)
    
    # 同一会社同一営業所の場合は増車と減車の同時アップロードを許可
    # 異なる会社・営業所の場合は増車と減車の同時アップロードを禁止
    if inc_files and dec_files:
        # 既に同一会社同一営業所のチェックは上で通過しているので、同時アップロードを許可
        pass

    company_name = next(iter(companies))
    office_name = next(iter(offices))
    office_disp, _, _ = parse_office_display(office_name)

    db_flat = load_db_raw()
    idx = find_company_index(db_flat, company_name, office_name)
    if idx < 0:
        return js_alert(f"会社・拠点がDBに見つかりません。会社={company_name} / 営業所={office_name}", status=404)

    old_counts = get_type_counts_from_row(db_flat["records"][idx])
    
    # ファイル別の減車数をチェック
    dec_n = len(dec_files)
    if dec_n:
        # 各ファイルの種別ごとに減車数をチェック
        car_type_counts = {}
        for car_type in dec_car_types:
            car_type_counts[car_type] = car_type_counts.get(car_type, 0) + 1
        
        for car_type, count in car_type_counts.items():
            if old_counts.get(car_type, 0) - count < 0:
                return js_alert(f"現在、{company_name} {office_name} の {car_type} が不足します（減車{count}台は不可）。", status=409)

    tpl_path = SRC / TPL_NAME
    if not tpl_path.exists():
        return js_alert(f"テンプレートが見つかりません: {tpl_path}", status=500)

    work = OUTPUT / f"work_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copyfile(tpl_path, work)

    new_counts = old_counts.copy()
    
    # ファイル別の増車処理
    if inc_files:
        car_type_counts = {}
        weight_class_counts = {}
        for i, car_type in enumerate(inc_car_types):
            car_type_counts[car_type] = car_type_counts.get(car_type, 0) + 1
            weight_class = inc_weight_classes[i]
            weight_class_counts[weight_class] = weight_class_counts.get(weight_class, 0) + 1
        
        # 種別ごとにDB更新
        for car_type, count in car_type_counts.items():
            new_counts = update_db_counts_inplace(db_flat, idx, car_type, +count)
        
        # 重量クラスごとにDB更新
        for weight_class, count in weight_class_counts.items():
            update_db_weight_inplace(db_flat, idx, weight_class, +count)
    
    # ファイル別の減車処理
    if dec_files:
        car_type_counts = {}
        weight_class_counts = {}
        for i, car_type in enumerate(dec_car_types):
            car_type_counts[car_type] = car_type_counts.get(car_type, 0) + 1
            weight_class = dec_weight_classes[i]
            weight_class_counts[weight_class] = weight_class_counts.get(weight_class, 0) + 1
        
        # 種別ごとにDB更新
        for car_type, count in car_type_counts.items():
            new_counts = update_db_counts_inplace(db_flat, idx, car_type, -count)
        
        # 重量クラスごとにDB更新
        for weight_class, count in weight_class_counts.items():
            update_db_weight_inplace(db_flat, idx, weight_class, -count)
    put_weight_totals_to_sheet(work, db_flat["records"][idx])

    put_old_new_counts(work, office_disp, old_counts, new_counts)

    modes = (["increase"] * len(inc_files)) + (["decrease"] * len(dec_files))
    
    ok, why = is_valid_xlsx(work)
    if not ok:
        return fail(f"出力ファイルが不正です（{why}）。サーバログを確認してください。")

    out_name = OUTPUT / f"事業計画変更届出_{office_disp or '不明営業所'}.xlsx"
    cnt = 2
    while out_name.exists():
        out_name = out_name.with_stem(out_name.stem + f"({cnt})")
        cnt += 1

    shutil.move(work, out_name)


    # C/Mブロック：1〜5件目はC列、6〜10件目はM列のみ不要語を削除
    apply_label_toggle_blocked(out_name, modes, sheet_index=0)
    # F列：最大積載量
    apply_max_loads(out_name, modes, inc_files, dec_files, sheet_index=0, dest_col="F")
    # 年式：1〜5件目はH列、6〜10件目はR列
    apply_model_years(out_name, modes, inc_files, dec_files, sheet_index=0)

    prefer = "increase" if len(inc_files) > 0 else "decrease"
    prepend_honshi_sheet(out_name, prefer)
    
    # C/Mの片側だけ不要語を削除 + F列に最大積載量

# XLSXのMIMEを明示
    return send_file(
        out_name,
        as_attachment=True,
        download_name=out_name.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if __name__=="__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
